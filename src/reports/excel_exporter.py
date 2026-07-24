# -*- coding: utf-8 -*-

"""
Exportación completa a Excel por participante.

SportsLabResearch-BreastCancer-Wellbeing-Analyzer
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.utils import get_column_letter

from src.analysis.blood_pressure import analyse as analyse_blood_pressure
from src.analysis.wellbeing import analyse as analyse_wellbeing
from src.cleaning.clinical_cleaning import clean_clinical_data
from src.config.variable_mapping import prepare_form_dataframe
from src.core.session import get_current_participant
from src.graphics.clinical_graphs import generate_clinical_graphs


PROJECT_NAME = "SportsLabResearch-BreastCancer-Wellbeing-Analyzer"

ROOT = Path(__file__).resolve().parents[2]
RESULTS_DIR = ROOT / "results"

VARIABLES = {
    "hr": ("Frecuencia cardiaca", "bpm"),
    "rmssd": ("RMSSD", "ms"),
    "ln_rmssd": ("LnRMSSD", "ln(ms)"),
    "sbp": ("Presión arterial sistólica", "mmHg"),
    "dbp": ("Presión arterial diastólica", "mmHg"),
    "spo2": ("Saturación de oxígeno", "%"),
    "sleep": ("Sueño", "puntos"),
    "mood": ("Estado de ánimo", "puntos"),
    "stress": ("Estrés", "puntos"),
    "fatigue": ("Fatiga", "puntos"),
    "upper_pain": ("Dolor superior", "puntos"),
    "lower_pain": ("Dolor inferior", "puntos"),
}


def safe_filename(value: object) -> str:
    text = str(value or "").strip()

    replacements = {
        " ": "_",
        "/": "-",
        "\\": "-",
        ":": "-",
        "*": "",
        "?": "",
        '"': "",
        "<": "",
        ">": "",
        "|": "",
    }

    for old, new in replacements.items():
        text = text.replace(old, new)

    return text or "participante"


def build_summary(
    participant: dict[str, Any],
    clinical_data: pd.DataFrame,
) -> pd.DataFrame:
    if "date" in clinical_data.columns:
        dates = pd.to_datetime(
            clinical_data["date"],
            errors="coerce",
            dayfirst=True,
        )
    else:
        dates = pd.Series(dtype="datetime64[ns]")

    first_date = (
        dates.min().strftime("%d/%m/%Y")
        if not dates.empty and dates.notna().any()
        else "No disponible"
    )

    last_date = (
        dates.max().strftime("%d/%m/%Y")
        if not dates.empty and dates.notna().any()
        else "No disponible"
    )

    rows = [
        ("Proyecto", PROJECT_NAME),
        (
            "Identificador",
            participant.get("participant_id", "No disponible"),
        ),
        (
            "Nombre",
            participant.get("name", "No disponible"),
        ),
        (
            "Sede",
            participant.get("site", "No disponible"),
        ),
        ("Registros analizados", len(clinical_data)),
        ("Primer registro", first_date),
        ("Último registro", last_date),
        (
            "Fecha de exportación",
            datetime.now().strftime("%d/%m/%Y %H:%M"),
        ),
    ]

    return pd.DataFrame(
        rows,
        columns=["Campo", "Valor"],
    )


def build_descriptive(
    clinical_data: pd.DataFrame,
) -> pd.DataFrame:
    rows = []

    for column, (label, unit) in VARIABLES.items():
        if column not in clinical_data.columns:
            continue

        values = pd.to_numeric(
            clinical_data[column],
            errors="coerce",
        ).dropna()

        if values.empty:
            continue

        rows.append(
            {
                "Variable": label,
                "Unidad": unit,
                "n": int(values.count()),
                "Media": values.mean(),
                "DE": values.std(),
                "Mediana": values.median(),
                "Mínimo": values.min(),
                "Máximo": values.max(),
            }
        )

    return pd.DataFrame(rows)


def build_data_quality(
    clinical_data: pd.DataFrame,
) -> pd.DataFrame:
    rows = []

    for column, (label, unit) in VARIABLES.items():
        if column not in clinical_data.columns:
            continue

        valid = int(clinical_data[column].notna().sum())
        missing = int(clinical_data[column].isna().sum())
        total = valid + missing

        availability = (
            valid / total * 100
            if total > 0
            else 0
        )

        rows.append(
            {
                "Variable": label,
                "Unidad": unit,
                "Registros válidos": valid,
                "Registros ausentes": missing,
                "Disponibilidad (%)": availability,
            }
        )

    return pd.DataFrame(rows)


def build_blood_pressure(
    clinical_data: pd.DataFrame,
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    try:
        results = analyse_blood_pressure(clinical_data)

        summary = pd.DataFrame(
            [
                {
                    "Registros totales": results["summary"].get("records"),
                    "PAS válidas": results["summary"].get("valid_sbp"),
                    "PAD válidas": results["summary"].get("valid_dbp"),
                    "PAS media": results["summary"].get("sbp_mean"),
                    "PAD media": results["summary"].get("dbp_mean"),
                    "Clasificación": results["summary"].get(
                        "classification"
                    ),
                    "Interpretación": results["summary"].get(
                        "interpretation"
                    ),
                }
            ]
        )

        descriptive = results.get(
            "descriptive",
            pd.DataFrame(),
        )

        alerts = results.get(
            "alerts",
            pd.DataFrame(),
        )

        if not isinstance(alerts, pd.DataFrame):
            alerts = pd.DataFrame(alerts)

        return summary, descriptive, alerts

    except Exception as exc:
        error = pd.DataFrame(
            [{"Error": str(exc)}]
        )

        return error, pd.DataFrame(), pd.DataFrame()


def build_wellbeing(
    clinical_data: pd.DataFrame,
) -> pd.DataFrame:
    try:
        results = analyse_wellbeing(clinical_data)

        if results is None:
            return pd.DataFrame()

        return results.copy()

    except Exception as exc:
        return pd.DataFrame(
            [{"Error": str(exc)}]
        )


def prepare_export_data(
    records: pd.DataFrame,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    original = records.copy()

    clinical = prepare_form_dataframe(
        records.copy()
    )

    clinical = clean_clinical_data(
        clinical
    )

    return original, clinical


def adjust_workbook(path: Path) -> None:
    workbook = load_workbook(path)

    header_fill = PatternFill(
        fill_type="solid",
        fgColor="1F4E78",
    )

    header_font = Font(
        color="FFFFFF",
        bold=True,
    )

    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )

        for column_cells in worksheet.columns:
            max_length = 0

            column_letter = get_column_letter(
                column_cells[0].column
            )

            for cell in column_cells:
                value = "" if cell.value is None else str(cell.value)
                max_length = max(
                    max_length,
                    len(value),
                )

                cell.alignment = Alignment(
                    vertical="top",
                    wrap_text=True,
                )

            worksheet.column_dimensions[
                column_letter
            ].width = min(
                max(max_length + 2, 12),
                45,
            )

        worksheet.auto_filter.ref = worksheet.dimensions

    workbook.save(path)



def add_graphs_to_workbook(
    workbook_path: Path,
    records: pd.DataFrame,
    participant_folder: Path,
) -> None:
    figures_folder = participant_folder / "figures"

    graph_paths = generate_clinical_graphs(
        records,
        output_dir=figures_folder,
    )

    if not graph_paths:
        return

    workbook = load_workbook(workbook_path)

    if "Graficos" in workbook.sheetnames:
        del workbook["Graficos"]

    worksheet = workbook.create_sheet(
        title="Graficos"
    )

    worksheet.sheet_view.showGridLines = False

    worksheet["A1"] = "GRAFICOS CLINICOS"
    worksheet["A1"].font = Font(
        bold=True,
        size=16,
    )

    worksheet["A2"] = (
        "Evoluci?n longitudinal de las variables "
        "correspondientes al intervalo seleccionado."
    )

    worksheet["A2"].alignment = Alignment(
        wrap_text=True,
        vertical="top",
    )

    graph_titles = {
        "presion_arterial.png":
            "Evoluci?n de la presi?n arterial",
        "frecuencia_cardiaca.png":
            "Evoluci?n de la frecuencia cardiaca",
        "rmssd.png":
            "Evoluci?n de RMSSD",
        "ln_rmssd.png":
            "Evoluci?n de LnRMSSD",
        "spo2.png":
            "Evoluci?n de la saturaci?n de ox?geno",
        "sueno.png":
            "Evoluci?n del sue?o",
        "estado_animo.png":
            "Evoluci?n del estado de ?nimo",
        "estres.png":
            "Evoluci?n del estr?s",
        "fatiga.png":
            "Evoluci?n de la fatiga",
        "dolor_superior.png":
            "Evoluci?n del dolor superior",
        "dolor_inferior.png":
            "Evoluci?n del dolor inferior",
    }

    current_row = 4

    for graph_path in graph_paths:
        title = graph_titles.get(
            graph_path.name,
            graph_path.stem.replace("_", " ").title(),
        )

        title_cell = worksheet.cell(
            row=current_row,
            column=1,
            value=title,
        )

        title_cell.font = Font(
            bold=True,
            size=12,
        )

        image = ExcelImage(
            str(graph_path)
        )

        image.width = 900
        image.height = 430

        image_anchor = f"A{current_row + 1}"

        worksheet.add_image(
            image,
            image_anchor,
        )

        current_row += 25

    worksheet.column_dimensions["A"].width = 120

    workbook.save(workbook_path)


def export_participant_excel(
    records: pd.DataFrame,
    output_dir: Path | str | None = None,
    include_graphs: bool = True,
) -> Path:
    if records is None or records.empty:
        raise ValueError(
            "No hay registros disponibles para exportar."
        )

    participant = get_current_participant()

    if not participant:
        participant = {
            "participant_id": "sin_id",
            "name": "Participante",
            "site": "No disponible",
        }

    original_data, clinical_data = prepare_export_data(
        records
    )

    if output_dir is None:
        participant_folder = (
            RESULTS_DIR
            / safe_filename(
                participant.get("site")
            )
            / (
                safe_filename(
                    participant.get("participant_id")
                )
                + "_"
                + safe_filename(
                    participant.get("name")
                )
            )
        )
    else:
        participant_folder = Path(output_dir)

    participant_folder.mkdir(
        parents=True,
        exist_ok=True,
    )

    filename = (
        "Resultados_"
        + safe_filename(
            participant.get("participant_id")
        )
        + "_"
        + safe_filename(
            participant.get("name")
        )
        + ".xlsx"
    )

    output_path = participant_folder / filename

    summary = build_summary(
        participant,
        clinical_data,
    )

    descriptive = build_descriptive(
        clinical_data
    )

    quality = build_data_quality(
        clinical_data
    )

    wellbeing = build_wellbeing(
        clinical_data
    )

    (
        pressure_summary,
        pressure_descriptive,
        pressure_alerts,
    ) = build_blood_pressure(
        clinical_data
    )

    with pd.ExcelWriter(
        output_path,
        engine="openpyxl",
    ) as writer:
        summary.to_excel(
            writer,
            sheet_name="Resumen",
            index=False,
        )

        original_data.to_excel(
            writer,
            sheet_name="Datos originales",
            index=False,
        )

        clinical_data.to_excel(
            writer,
            sheet_name="Datos depurados",
            index=False,
        )

        descriptive.to_excel(
            writer,
            sheet_name="Descriptivos",
            index=False,
        )

        wellbeing.to_excel(
            writer,
            sheet_name="Bienestar",
            index=False,
        )

        pressure_summary.to_excel(
            writer,
            sheet_name="PA resumen",
            index=False,
        )

        pressure_descriptive.to_excel(
            writer,
            sheet_name="PA descriptivos",
            index=False,
        )

        pressure_alerts.to_excel(
            writer,
            sheet_name="PA alertas",
            index=False,
        )

        quality.to_excel(
            writer,
            sheet_name="Calidad datos",
            index=False,
        )

    adjust_workbook(output_path)

    if include_graphs:
        add_graphs_to_workbook(
            output_path,
            records,
            participant_folder,
        )

    return output_path
